#!/usr/bin/env python3
"""
Generate sample data files for datui testing.

This script generates various CSV, Parquet, IPC/Arrow, Avro, and Excel files:
- Different data types
- Missing data (nulls)
- Empty tables
- Quoted and unquoted strings
- Files for grouping operations
- Files for aggregate calculations
- Large and small datasets
- Error case testing
- Pivot and Melt reshape testing (long-format for pivot, wide-format for melt)
- Correlation matrix demo (100k rows, 10 numeric columns with varying correlations)

Uses Polars for most formats; fastavro for Avro; openpyxl for Excel.
"""

import os
import sys
from pathlib import Path
import polars as pl
import numpy as np
from datetime import date, datetime, timedelta
import random
import gzip

# Optional deps for extra formats (fail gracefully if missing)
try:
    import fastavro
except ImportError:
    fastavro = None
try:
    import openpyxl
except ImportError:
    openpyxl = None

# Output directory
OUTPUT_DIR = Path(__file__).parent.parent / "tests" / "sample-data"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

def generate_people_data():
    """Generate a people database with cities, states, etc. for grouping."""
    np.random.seed(42)
    random.seed(42)
    
    cities = ["Springfield", "Riverside", "Franklin", "Greenville", "Bristol", 
              "Madison", "Clinton", "Marion", "Georgetown", "Salem"]
    states = ["CA", "NY", "TX", "FL", "IL", "PA", "OH", "GA", "NC", "MI"]
    departments = ["Engineering", "Sales", "Marketing", "HR", "Finance", "Operations"]
    job_titles = ["Manager", "Senior", "Junior", "Lead", "Director", "Analyst"]
    
    n = 1000
    data = {
        "id": list(range(1, n + 1)),
        "first_name": [f"Person{i}" for i in range(1, n + 1)],
        "last_name": [f"Lastname{i}" for i in range(1, n + 1)],
        "age": np.random.randint(22, 65, n).tolist(),
        "city": np.random.choice(cities, n).tolist(),
        "state": np.random.choice(states, n).tolist(),
        "department": np.random.choice(departments, n).tolist(),
        "job_title": np.random.choice(job_titles, n).tolist(),
        "salary": np.random.randint(40000, 150000, n).tolist(),
        "start_date": [(datetime(2020, 1, 1) + timedelta(days=random.randint(0, 1460))).date() for _ in range(n)],
        "active": np.random.choice([True, False], n, p=[0.8, 0.2]).tolist(),
    }
    
    df = pl.DataFrame(data)
    
    # Add some nulls - create a mask for each column
    null_count = int(n * 0.1)
    for col in ["city", "department", "salary"]:
        null_indices = set(np.random.choice(n, size=null_count, replace=False))
        mask = [i in null_indices for i in range(n)]
        if col == "salary":
            df = df.with_columns(
                pl.when(pl.Series("mask", mask))
                .then(None)
                .otherwise(pl.col(col))
                .alias(col)
            )
        else:
            df = df.with_columns(
                pl.when(pl.Series("mask", mask))
                .then(None)
                .otherwise(pl.col(col))
                .alias(col)
            )
    
    return df

def generate_sales_data():
    """Generate sales data for aggregate calculations."""
    np.random.seed(43)
    random.seed(43)
    
    products = ["Widget A", "Widget B", "Widget C", "Gadget X", "Gadget Y", "Tool 1", "Tool 2"]
    regions = ["North", "South", "East", "West", "Central"]
    
    n = 5000
    data = {
        "date": [(datetime(2023, 1, 1) + timedelta(days=random.randint(0, 730))).date() for _ in range(n)],
        "product": np.random.choice(products, n).tolist(),
        "region": np.random.choice(regions, n).tolist(),
        "quantity": np.random.randint(1, 100, n).tolist(),
        "unit_price": [round(random.uniform(10.0, 500.0), 2) for _ in range(n)],
        "discount": [round(random.uniform(0.0, 0.3), 2) for _ in range(n)],
    }
    
    df = pl.DataFrame(data)
    df = df.with_columns(
        (pl.col("quantity") * pl.col("unit_price") * (1 - pl.col("discount"))).alias("total")
    )
    
    # Add some nulls
    null_count = int(n * 0.05)
    for col in ["quantity", "unit_price", "discount"]:
        null_indices = set(np.random.choice(n, size=null_count, replace=False))
        mask = [i in null_indices for i in range(n)]
        df = df.with_columns(
            pl.when(pl.Series("mask", mask))
            .then(None)
            .otherwise(pl.col(col))
            .alias(col)
        )
    
    # Recalculate total where we have nulls
    df = df.with_columns(
        pl.when(pl.col("quantity").is_null() | pl.col("unit_price").is_null() | pl.col("discount").is_null())
        .then(None)
        .otherwise(pl.col("quantity") * pl.col("unit_price") * (1 - pl.col("discount")))
        .alias("total")
    )
    
    return df

def generate_mixed_types():
    """Generate data with various types including nulls."""
    np.random.seed(44)
    
    n = 200
    data = {
        "id": list(range(1, n + 1)),
        "integer_col": np.random.randint(-100, 100, n).tolist(),
        "float_col": [round(random.uniform(-50.0, 50.0), 3) for _ in range(n)],
        "string_col": [f"text_{i}" for i in range(n)],
        "boolean_col": np.random.choice([True, False], n).tolist(),
        "date_col": [(datetime(2020, 1, 1) + timedelta(days=i)).date() for i in range(n)],
    }
    
    df = pl.DataFrame(data)
    
    # Add nulls to various columns
    null_count = int(n * 0.15)
    for col in ["integer_col", "float_col", "string_col", "boolean_col", "date_col"]:
        null_indices = set(np.random.choice(n, size=null_count, replace=False))
        mask = [i in null_indices for i in range(n)]
        df = df.with_columns(
            pl.when(pl.Series("mask", mask))
            .then(None)
            .otherwise(pl.col(col))
            .alias(col)
        )
    
    return df

def generate_quoted_strings():
    """Generate CSV with quoted strings containing commas, newlines, etc."""
    data = {
        "id": [1, 2, 3, 4, 5],
        "name": [
            "Normal Name",
            "Name, with comma",
            "Name\nwith newline",
            'Name "with quotes"',
            "Name, with\nmultiple, issues",
        ],
        "description": [
            "Simple description",
            "Description, with comma, and more",
            "Description\nwith\nnewlines",
            'Description with "quotes" and, commas',
            "Complex: has, commas\nand newlines\nand \"quotes\"",
        ],
        "value": [10, 20, 30, 40, 50],
    }
    
    df = pl.DataFrame(data)
    return df

def generate_empty_table():
    """Generate an empty table with schema."""
    df = pl.DataFrame({
        "id": pl.Series([], dtype=pl.Int64),
        "name": pl.Series([], dtype=pl.Utf8),
        "value": pl.Series([], dtype=pl.Float64),
        "date": pl.Series([], dtype=pl.Date),
    })
    return df

def generate_single_row():
    """Generate a table with a single row."""
    data = {
        "id": [1],
        "name": ["Single Row"],
        "value": [42],
        "date": [datetime(2024, 1, 1).date()],
    }
    df = pl.DataFrame(data)
    return df

def generate_large_dataset():
    """Generate a large dataset for performance testing with various distributions."""
    np.random.seed(45)
    random.seed(45)
    
    n = 1000000
    
    # Generate distributions with various characteristics
    # Preserve distribution characteristics for proper detection testing
    
    # Normal distribution (can have negative values) - keep natural scale
    normal_data = np.random.normal(loc=0.0, scale=1.0, size=n)
    
    # LogNormal distribution (positive values) - keep natural scale
    lognormal_data = np.random.lognormal(mean=0.0, sigma=1.0, size=n)
    
    # Uniform distribution - keep in [0, 1] (natural for uniform)
    uniform_data = np.random.uniform(0.0, 1.0, n)
    
    # Power Law distribution (positive values) - keep natural scale
    # Generate using inverse transform: x = xmin * (1 - u)^(-1/(alpha-1))
    # where u is uniform [0,1] and alpha > 1
    alpha = 2.5
    xmin = 1.0  # Start from 1.0 for better power-law characteristics
    powerlaw_data = xmin * np.power(1.0 - np.random.uniform(0.0, 1.0, n), -1.0 / (alpha - 1.0))
    
    # Exponential distribution (positive values) - keep natural scale
    lambda_param = 2.0
    exponential_data = np.random.exponential(scale=1.0/lambda_param, size=n)
    
    # Beta distribution - naturally in [0, 1], keep as is
    beta_data = np.random.beta(a=2.0, b=5.0, size=n)
    
    # Gamma distribution (positive values) - keep natural scale
    shape = 2.0
    scale = 0.5
    gamma_data = np.random.gamma(shape=shape, scale=scale, size=n)
    
    # Chi-squared distribution (non-negative) - keep natural scale
    df = 5.0
    chisq_data = np.random.chisquare(df=df, size=n)
    
    # Student's t distribution (can have negative values) - keep natural scale
    t_df = 5.0
    t_data = np.random.standard_t(df=t_df, size=n)
    
    # Poisson distribution (non-negative integers) - KEEP AS INTEGERS
    lambda_poisson = 5.0
    poisson_data = np.random.poisson(lam=lambda_poisson, size=n).astype(int)
    
    # Bernoulli distribution - KEEP AS BINARY INTEGERS [0, 1]
    p_bernoulli = 0.3
    bernoulli_data = np.random.binomial(n=1, p=p_bernoulli, size=n).astype(int)
    
    # Binomial distribution (non-negative integers) - KEEP AS INTEGERS
    n_binomial = 20
    p_binomial = 0.4
    binomial_data = np.random.binomial(n=n_binomial, p=p_binomial, size=n).astype(int)
    
    # Geometric distribution (non-negative integers) - KEEP AS INTEGERS
    p_geometric = 0.3
    geometric_data = np.random.geometric(p=p_geometric, size=n).astype(int)
    
    # Weibull distribution (positive values) - keep natural scale
    weibull_shape = 2.0
    weibull_scale = 1.0
    weibull_data = weibull_scale * np.power(-np.log(np.random.uniform(0.001, 1.0, n)), 1.0 / weibull_shape)
    
    # Generate all 2-letter combinations (AA, AB, ..., ZZ) = 26*26 = 676 categories
    categories = [f"{chr(65+i)}{chr(65+j)}" for i in range(26) for j in range(26)]
    num_categories = len(categories)

    # Generate power-law distributed categories
    power_law_values = np.random.power(a=1.5, size=n)   # 1.5 alpha for power law
    category_indices = (power_law_values * num_categories).astype(int)
    category_indices = np.clip(category_indices, 0, num_categories - 1)
    category_data = [categories[idx] for idx in category_indices]
    
    data = {
        "id": list(range(1, n + 1)),
        "category": category_data,
        "value1": np.random.randint(0, 1000, n).tolist(),
        "value2": [round(random.uniform(0.0, 100.0), 2) for _ in range(n)],
        "value3": np.random.choice([True, False], n).tolist(),
        "timestamp": [datetime(2024, 1, 1) + timedelta(seconds=i) for i in range(n)],
        # Distribution columns
        # Continuous distributions: round to 6 decimal places
        "dist_normal": [round(float(x), 6) for x in normal_data],
        "dist_lognormal": [round(float(x), 6) for x in lognormal_data],
        "dist_uniform": [round(float(x), 6) for x in uniform_data],
        "dist_powerlaw": [round(float(x), 6) for x in powerlaw_data],
        "dist_exponential": [round(float(x), 6) for x in exponential_data],
        "dist_beta": [round(float(x), 6) for x in beta_data],
        "dist_gamma": [round(float(x), 6) for x in gamma_data],
        "dist_chisquared": [round(float(x), 6) for x in chisq_data],
        "dist_students_t": [round(float(x), 6) for x in t_data],
        "dist_weibull": [round(float(x), 6) for x in weibull_data],
        # Discrete distributions: keep as integers (no rounding needed, but convert to list)
        "dist_poisson": poisson_data.tolist(),
        "dist_bernoulli": bernoulli_data.tolist(),
        "dist_binomial": binomial_data.tolist(),
        "dist_geometric": geometric_data.tolist(),
    }
    
    df = pl.DataFrame(data)
    return df

def generate_error_cases():
    """Generate files that test error cases."""
    error_cases = {}
    
    # Case 1: Inconsistent types in column (convert all to strings to simulate mixed types)
    error_cases["inconsistent_types"] = pl.DataFrame({
        "id": pl.Series(["1", "2", "3", "not_a_number", "5"], dtype=pl.Utf8),  # All strings, but some look like numbers
        "value": [10, 20, 30, 40, 50],
    })
    
    # Case 2: Very long strings
    error_cases["long_strings"] = pl.DataFrame({
        "id": list(range(1, 11)),
        "long_text": ["A" * 1000] * 10,
    })
    
    # Case 3: Special characters
    error_cases["special_chars"] = pl.DataFrame({
        "id": list(range(1, 6)),
        "text": ["\x00", "\t", "\n", "\r", "\\"],
        "unicode": ["αβγ", "🚀", "中文", "العربية", "русский"],
    })
    
    return error_cases


# -----------------------------------------------------------------------------
# Pivot / Melt testing (see plans/pivot-melt-plan.md)
# -----------------------------------------------------------------------------


def generate_pivot_long():
    """
    Long-format data for Pivot tab testing.

    Schema: id, date, key, value (float).
    - Multiple rows per (id, date) with distinct keys "A", "B", "C".
    - Some (id, date, key) duplicates to exercise aggregation (last, first, min, max, etc.).
    - Deterministic (seed 50) for reproducible tests.
    """
    np.random.seed(50)
    random.seed(50)

    keys = ["A", "B", "C"]
    n_groups = 40
    base_date = datetime(2024, 1, 1).date()

    rows = []
    for g in range(n_groups):
        uid = (g % 20) + 1
        d = base_date + timedelta(days=g % 31)
        for k in keys:
            rows.append({"id": uid, "date": d, "key": k, "value": round(random.uniform(10.0, 100.0), 2)})

    # Add duplicates for aggregation tests: same (id, date, key), different value
    n_dup = 24
    for _ in range(n_dup):
        r = random.choice(rows)
        rows.append({
            "id": r["id"],
            "date": r["date"],
            "key": r["key"],
            "value": round(random.uniform(10.0, 100.0), 2),
        })

    return pl.DataFrame(rows)


def generate_pivot_long_string():
    """
    Long-format data with string value column for Pivot + first/last aggregation.

    Schema: id, date, key, value (str). Same structure as pivot_long; value is
    "low", "mid", or "high" so only first/last aggregation is meaningful.
    """
    np.random.seed(51)
    random.seed(51)

    keys = ["X", "Y", "Z"]
    labels = ["low", "mid", "high"]
    n_groups = 30
    base_date = datetime(2024, 1, 1).date()

    rows = []
    for g in range(n_groups):
        uid = (g % 15) + 1
        d = base_date + timedelta(days=g % 28)
        for k in keys:
            rows.append({"id": uid, "date": d, "key": k, "value": random.choice(labels)})

    return pl.DataFrame(rows)


def generate_melt_wide():
    """
    Wide-format data for Melt tab testing.

    Schema: id, date, Q1_2024..Q4_2024, metric_foo, metric_bar, label.
    - Pattern-friendly names for regex tests (Q[1-4]_2024, metric_*).
    - Mix of numeric and string (label) for "by type" tests.
    """
    np.random.seed(52)
    random.seed(52)

    n = 80
    base_date = datetime(2024, 1, 1).date()
    data = {
        "id": list(range(1, n + 1)),
        "date": [base_date + timedelta(days=random.randint(0, 365)) for _ in range(n)],
        "Q1_2024": [round(random.uniform(0, 100), 2) for _ in range(n)],
        "Q2_2024": [round(random.uniform(0, 100), 2) for _ in range(n)],
        "Q3_2024": [round(random.uniform(0, 100), 2) for _ in range(n)],
        "Q4_2024": [round(random.uniform(0, 100), 2) for _ in range(n)],
        "metric_foo": [round(random.uniform(0, 50), 2) for _ in range(n)],
        "metric_bar": [round(random.uniform(0, 50), 2) for _ in range(n)],
        "label": random.choices(["alpha", "beta", "gamma"], k=n),
    }
    return pl.DataFrame(data)


def generate_melt_wide_many():
    """
    Wide-format data with many value columns for Melt "all except index" / pattern stress.

    Schema: id, date, col_1, col_2, ..., col_50. All numeric except id/date.
    """
    np.random.seed(53)
    random.seed(53)

    n = 60
    n_cols = 50
    base_date = datetime(2024, 1, 1).date()

    data = {
        "id": list(range(1, n + 1)),
        "date": [base_date + timedelta(days=random.randint(0, 200)) for _ in range(n)],
    }
    for i in range(1, n_cols + 1):
        data[f"col_{i}"] = [round(random.uniform(0, 100), 2) for _ in range(n)]

    return pl.DataFrame(data)


def generate_charting_demo():
    """
    Generate daily time-series data for chart view demos and testing.

    One row per day for 10 years (2015-01-01 through 2024-12-31). Columns:
    - date: sequential daily dates
    - day_of_week: Mon, Tue, ..., Sun (categorical)
    - stock_market: fictitious index (random walk with drift/volatility)
    - high_temp: daily high (seasonal + noise)
    - 20d_avg_high_temp: 20-day rolling average of high_temp
    - customer_count: integer (seasonal + weekday + noise)
    - shark_sightings: integer daily count (low with occasional spikes)
    """
    np.random.seed(55)
    random.seed(55)

    base = datetime(2015, 1, 1).date()
    days = (datetime(2024, 12, 31).date() - base).days + 1
    dates = [base + timedelta(days=i) for i in range(days)]
    day_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

    # day_of_week from date (2015-01-01 is Thursday → weekday 3)
    day_of_week = [
        day_names[(base.weekday() + i) % 7] for i in range(days)
    ]

    # Stock market: random walk with slight upward drift and volatility
    walk = np.zeros(days)
    walk[0] = 1000.0
    for i in range(1, days):
        walk[i] = walk[i - 1] + np.random.normal(0.5, 15.0)
    stock_market = [round(float(x), 2) for x in walk]

    # High temp: seasonal sine + noise (roughly 30–100 °F)
    t = np.arange(days, dtype=float)
    seasonal = 65.0 + 25.0 * np.sin(2 * np.pi * t / 365.25 - 1.6)
    noise = np.random.normal(0, 5.0, days)
    high_temp = [round(float(np.clip(seasonal[i] + noise[i], 25, 105)), 1) for i in range(days)]

    # 20-day rolling average of high_temp
    high_temp_arr = np.array(high_temp, dtype=float)
    avg_20d = np.zeros(days)
    for i in range(days):
        lo = max(0, i - 19)
        avg_20d[i] = high_temp_arr[lo : i + 1].mean()
    avg_20d_list = [round(float(x), 1) for x in avg_20d]

    # Customer count: base + weekday effect + seasonal + noise (non-negative int)
    weekday_effect = np.array([0, 0, 0, 0, 10, 25, 15])  # Fri/Sat/Sun higher
    base_cust = 500
    cust = np.zeros(days)
    for i in range(days):
        wd = (base.weekday() + i) % 7
        seasonal_cust = 80 * np.sin(2 * np.pi * i / 365.25)
        cust[i] = base_cust + weekday_effect[wd] + seasonal_cust + np.random.normal(0, 30)
    customer_count = [max(0, int(round(x))) for x in cust]

    # Shark sightings: low counts, occasional spikes (Poisson-like with rare spikes)
    lam = np.ones(days) * 0.3
    for _ in range(12):
        idx = random.randint(0, days - 1)
        lam[idx] = 8.0 + random.uniform(0, 5)
    shark_sightings = [np.random.poisson(l) for l in lam]

    data = {
        "date": dates,
        "day_of_week": day_of_week,
        "stock_market": stock_market,
        "high_temp": high_temp,
        "20d_avg_high_temp": avg_20d_list,
        "customer_count": customer_count,
        "shark_sightings": shark_sightings,
    }
    return pl.DataFrame(data)


def generate_correlation_matrix_data():
    """
    Generate numeric data with designed pairwise correlations for correlation matrix demos.

    Produces 100_000 rows and 10 numeric columns with realistic names. Correlations
    are chosen to span the full range (strong negative to strong positive) so the
    correlation matrix heatmap uses the full color scale.

    Columns: revenue, profit, operating_cost, margin_pct, unit_volume, price_index,
             growth_rate, market_share, roi, cash_flow
    """
    np.random.seed(54)
    random.seed(54)

    n = 100_000
    col_names = [
        "revenue",
        "profit",
        "operating_cost",
        "margin_pct",
        "unit_volume",
        "price_index",
        "growth_rate",
        "market_share",
        "roi",
        "cash_flow",
    ]
    k = len(col_names)

    # Target correlation matrix (symmetric, 1 on diagonal). Order matches col_names.
    # Designed for variety: strong +/-, moderate +/-, weak +/-, near zero.
    R = np.array(
        [
            [1.00, 0.92, 0.88, 0.45, 0.72, 0.15, 0.08, -0.05, 0.68, 0.85],   # revenue
            [0.92, 1.00, 0.78, 0.82, 0.58, 0.22, 0.12, 0.02, 0.90, 0.88],   # profit
            [0.88, 0.78, 1.00, -0.75, 0.65, -0.10, -0.02, -0.08, 0.52, 0.62], # operating_cost
            [0.45, 0.82, -0.75, 1.00, 0.20, 0.55, 0.18, 0.25, 0.78, 0.70],   # margin_pct
            [0.72, 0.58, 0.65, 0.20, 1.00, -0.35, 0.30, 0.40, 0.35, 0.48],   # unit_volume
            [0.15, 0.22, -0.10, 0.55, -0.35, 1.00, 0.05, 0.12, 0.28, 0.18], # price_index
            [0.08, 0.12, -0.02, 0.18, 0.30, 0.05, 1.00, 0.42, 0.15, 0.10],   # growth_rate
            [-0.05, 0.02, -0.08, 0.25, 0.40, 0.12, 0.42, 1.00, 0.08, 0.02],  # market_share
            [0.68, 0.90, 0.52, 0.78, 0.35, 0.28, 0.15, 0.08, 1.00, 0.82],   # roi
            [0.85, 0.88, 0.62, 0.70, 0.48, 0.18, 0.10, 0.02, 0.82, 1.00],   # cash_flow
        ],
        dtype=np.float64,
    )

    # Standard deviations (scale each column to realistic ranges)
    scales = np.array([2.5e6, 4e5, 1.8e6, 8.0, 1.2e4, 15.0, 0.12, 5.0, 0.25, 3e5])
    # Covariance = diag(scales) @ R @ diag(scales)
    cov = np.outer(scales, scales) * R

    # Ensure covariance is positive definite (numerical safety)
    cov = (cov + cov.T) / 2
    min_eig = np.min(np.linalg.eigvalsh(cov))
    if min_eig < 1e-6:
        cov += (1e-6 - min_eig) * np.eye(k)

    mean = np.array(
        [1e7, 1.5e6, 6e6, 22.0, 5e4, 100.0, 0.05, 12.0, 0.15, 1e6],
        dtype=np.float64,
    )

    raw = np.random.multivariate_normal(mean, cov, size=n)

    # Clip to plausible non-negative ranges where needed (e.g. revenue, profit, %)
    raw[:, 0] = np.clip(raw[:, 0], 1e5, None)   # revenue
    raw[:, 1] = np.clip(raw[:, 1], -1e6, None)  # profit can be negative
    raw[:, 2] = np.clip(raw[:, 2], 1e4, None)   # operating_cost
    raw[:, 3] = np.clip(raw[:, 3], 0.1, 60.0)  # margin_pct
    raw[:, 4] = np.clip(raw[:, 4], 100, None)   # unit_volume
    raw[:, 5] = np.clip(raw[:, 5], 50, 200)    # price_index
    raw[:, 6] = np.clip(raw[:, 6], -0.5, 0.8)  # growth_rate
    raw[:, 7] = np.clip(raw[:, 7], 0, 40)      # market_share
    raw[:, 8] = np.clip(raw[:, 8], -0.2, 0.6)  # roi
    raw[:, 9] = np.clip(raw[:, 9], -5e5, None) # cash_flow can be negative

    data = {col_names[i]: [round(float(x), 4) for x in raw[:, i]] for i in range(k)}
    return pl.DataFrame(data)


def save_csv(df, filename, **kwargs):
    """Save DataFrame as CSV, compressed with gzip."""
    # Remove .csv extension if present, we'll add .csv.gz
    base_name = filename.replace('.csv', '')
    filepath = OUTPUT_DIR / f"{base_name}.csv.gz"
    
    # Write to temporary file first, then compress
    temp_path = OUTPUT_DIR / f"{base_name}.csv.tmp"
    df.write_csv(temp_path, **kwargs)
    
    # Compress the CSV file
    with open(temp_path, 'rb') as f_in:
        with gzip.open(filepath, 'wb', compresslevel=6) as f_out:
            f_out.writelines(f_in)
    
    # Remove temporary file
    temp_path.unlink()
    
    print(f"Generated: {filepath}")

def save_parquet(df, filename):
    """Save DataFrame as Parquet."""
    filepath = OUTPUT_DIR / filename
    df.write_parquet(filepath)
    print(f"Generated: {filepath}")


def save_ipc(df, filename):
    """Save DataFrame as Arrow IPC / Feather (e.g. .arrow, .ipc)."""
    base = filename.replace(".arrow", "").replace(".ipc", "")
    filepath = OUTPUT_DIR / f"{base}.arrow"
    df.write_ipc(filepath)
    print(f"Generated: {filepath}")


def _polars_dtype_to_avro(dtype):
    """Map Polars dtype to Avro schema (nullable union). Date/Datetime use logical types."""
    if dtype == pl.Int64:
        return ["null", "long"]
    if dtype == pl.Float64:
        return ["null", "double"]
    if dtype == pl.Utf8:
        return ["null", "string"]
    if dtype == pl.Boolean:
        return ["null", "boolean"]
    if dtype == pl.Date:
        return ["null", {"type": "int", "logicalType": "date"}]
    if dtype == pl.Datetime("us") or dtype == pl.Datetime("ms"):
        return ["null", {"type": "long", "logicalType": "timestamp-micros"}]
    # fallback
    return ["null", "string"]


def save_avro(df, filename):
    """Save DataFrame as Avro (requires fastavro)."""
    if fastavro is None:
        print("Skipping Avro (fastavro not installed):", filename)
        return
    filepath = OUTPUT_DIR / filename
    fields = []
    for name in df.columns:
        dtype = df.schema[name]
        avro_type = _polars_dtype_to_avro(dtype)
        fields.append({"name": name, "type": avro_type})
    schema = {"type": "record", "name": "Record", "fields": fields}
    parsed = fastavro.parse_schema(schema)

    epoch_date = date(1970, 1, 1)

    def row_to_record(row):
        record = {}
        for i, name in enumerate(df.columns):
            val = row[i]
            dtype = df.schema[name]
            if val is None:
                record[name] = None
            elif dtype == pl.Date and isinstance(val, date):
                record[name] = (val - epoch_date).days
            elif dtype in (pl.Datetime("us"), pl.Datetime("ms")) and hasattr(val, "timestamp"):
                record[name] = int(val.timestamp() * 1_000_000)
            elif hasattr(val, "isoformat"):
                record[name] = val.isoformat()
            else:
                record[name] = val
        return record

    records = [row_to_record(row) for row in df.iter_rows()]
    with open(filepath, "wb") as out:
        fastavro.writer(out, parsed, records, codec="deflate")
    print(f"Generated: {filepath}")


def save_excel(df, filename):
    """Save DataFrame as Excel .xlsx (requires openpyxl)."""
    if openpyxl is None:
        print("Skipping Excel (openpyxl not installed):", filename)
        return
    filepath = OUTPUT_DIR / filename
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        return
    for c, name in enumerate(df.columns, 1):
        ws.cell(row=1, column=c, value=name)
    for r in range(df.height):
        row = df.row(r)
        for c, val in enumerate(row, 1):
            if hasattr(val, "isoformat"):
                val = val.isoformat()
            ws.cell(row=r + 2, column=c, value=val)
    wb.save(filepath)
    print(f"Generated: {filepath}")


def main():
    print("Generating sample data files...")
    print(f"Output directory: {OUTPUT_DIR}")
    
    # People data for grouping
    print("\n1. Generating people data...")
    people_df = generate_people_data()
    save_csv(people_df, "people.csv")
    save_parquet(people_df, "people.parquet")
    save_ipc(people_df, "people.arrow")
    save_avro(people_df, "people.avro")
    save_excel(people_df, "people.xlsx")

    # Sales data for aggregates
    print("\n2. Generating sales data...")
    sales_df = generate_sales_data()
    save_csv(sales_df, "sales.csv")
    save_parquet(sales_df, "sales.parquet")
    save_ipc(sales_df, "sales.arrow")
    save_avro(sales_df, "sales.avro")
    save_excel(sales_df, "sales.xlsx")

    # Mixed types
    print("\n3. Generating mixed types data...")
    mixed_df = generate_mixed_types()
    save_csv(mixed_df, "mixed_types.csv")
    save_parquet(mixed_df, "mixed_types.parquet")
    save_ipc(mixed_df, "mixed_types.arrow")
    save_avro(mixed_df, "mixed_types.avro")
    save_excel(mixed_df, "mixed_types.xlsx")
    
    # Generate a small uncompressed CSV for testing (3 columns, good coverage)
    print("\n3a. Generating small test CSV (uncompressed)...")
    test_df = generate_mixed_types()  # Reuse mixed_types as it has good coverage
    # Save uncompressed version
    test_filepath = OUTPUT_DIR / "3-sfd-header.csv"
    test_df.write_csv(test_filepath)
    print(f"Generated: {test_filepath}")
    
    # Quoted strings
    print("\n4. Generating quoted strings data...")
    quoted_df = generate_quoted_strings()
    save_csv(quoted_df, "quoted_strings.csv")
    # For unquoted, we'll just save without special quoting (Polars handles this)
    save_csv(quoted_df, "unquoted_strings.csv")
    
    # Empty table
    print("\n5. Generating empty table...")
    empty_df = generate_empty_table()
    save_csv(empty_df, "empty.csv")
    save_parquet(empty_df, "empty.parquet")
    save_ipc(empty_df, "empty.arrow")
    save_avro(empty_df, "empty.avro")
    save_excel(empty_df, "empty.xlsx")

    # Single row
    print("\n6. Generating single row table...")
    single_df = generate_single_row()
    save_csv(single_df, "single_row.csv")
    save_parquet(single_df, "single_row.parquet")
    save_ipc(single_df, "single_row.arrow")
    save_avro(single_df, "single_row.avro")
    save_excel(single_df, "single_row.xlsx")
    
    # Large dataset
    print("\n7. Generating large dataset...")
    large_df = generate_large_dataset()
    save_csv(large_df, "large_dataset.csv")
    save_parquet(large_df, "large_dataset.parquet")
    
    # Error cases
    print("\n8. Generating error case files...")
    error_cases = generate_error_cases()
    for name, df in error_cases.items():
        save_csv(df, f"error_{name}.csv")
        # Skip parquet for inconsistent_types as it can't handle mixed types
        if name != "inconsistent_types":
            save_parquet(df, f"error_{name}.parquet")

    # Pivot and Melt testing
    print("\n9. Generating pivot and melt testing data...")
    pivot_long_df = generate_pivot_long()
    save_csv(pivot_long_df, "pivot_long.csv")
    save_parquet(pivot_long_df, "pivot_long.parquet")
    save_ipc(pivot_long_df, "pivot_long.arrow")
    save_avro(pivot_long_df, "pivot_long.avro")
    save_excel(pivot_long_df, "pivot_long.xlsx")
    pivot_long_string_df = generate_pivot_long_string()
    save_csv(pivot_long_string_df, "pivot_long_string.csv")
    save_parquet(pivot_long_string_df, "pivot_long_string.parquet")
    melt_wide_df = generate_melt_wide()
    save_csv(melt_wide_df, "melt_wide.csv")
    save_parquet(melt_wide_df, "melt_wide.parquet")
    save_ipc(melt_wide_df, "melt_wide.arrow")
    save_avro(melt_wide_df, "melt_wide.avro")
    save_excel(melt_wide_df, "melt_wide.xlsx")
    melt_wide_many_df = generate_melt_wide_many()
    save_csv(melt_wide_many_df, "melt_wide_many.csv")
    save_parquet(melt_wide_many_df, "melt_wide_many.parquet")

    # Charting demo (10 years daily time series)
    print("\n10. Generating charting demo data...")
    chart_df = generate_charting_demo()
    save_parquet(chart_df, "charting_demo.parquet")

    # Correlation matrix demo (Parquet only: 100k rows, 10 numeric columns)
    print("\n11. Generating correlation matrix demo data...")
    corr_df = generate_correlation_matrix_data()
    save_parquet(corr_df, "correlation_matrix_demo.parquet")

    print("\n✅ Sample data generation complete!")

if __name__ == "__main__":
    main()
